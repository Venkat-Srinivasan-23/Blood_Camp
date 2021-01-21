# SCRIPT NAME : Blood-Camp (MINI_PROJECT)
# DESCRIPTION : THIS SCRIPT ENABLES USERS TO GAVE A STRUCTURED INVENTORY OF BLOOD GROUP WHITH THE SPECIFIED REGION.
# INPUTS : FOR NOW SCHOOL RECORDS BEEN GIVEN. (We can also add any records).
# OUTPUT : AN EXCEL SHEET BEEN CREATED AND THE FILTERED OUTPUT IS BEEN PRINTED IN A STRUCTURED WAY.
# VERSION : 1
# AUTHOR : KRISHNAKUMAR

# libraries to install
import os
import glob
from pprint import pprint 
from openpyxl import Workbook
from openpyxl.styles import Font


def collect(sheet, filename, check, loc_check):
    file_open = open(filename)
    file_lines = file_open.readlines()

    count = sheet.max_row;

    name = []
    section = []
    roll_number = []
    ph_no = []
    blood_grp = []
    age = []
    location = []
    
    for lines in file_lines:
        if loc_check in lines:
            if check in lines:
                split_data = lines.split(' ')
                name.append(split_data[0])
                section.append(split_data[1])
                roll_number.append(split_data[2])
                ph_no.append(split_data[3])
                blood_grp.append(split_data[5])
                age.append(split_data[6])
                location.append(split_data[4])
                value = (name, blood_grp, ph_no, age, roll_number, location)
                #pprint(value)

                
    # to write in excel
    name_count = 0;
    for names in name:
        name_count = name_count + 1;
        student_name_cell = sheet.cell(row=count + name_count, column=1)
        student_name_cell.value = names

    section_count = 0;
    for sections in section:
        section_count = section_count + 1;
        class_cell = sheet.cell(row=count + section_count, column=2)
        class_cell.value = sections

    roll_count = 0;   
    for roll_numbers in roll_number:
        roll_count = roll_count + 1;
        roll_number_cell = sheet.cell(row=count + roll_count, column=3)
        roll_number_cell.value = roll_numbers

    ph_no_count = 0;
    for ph_nos in ph_no:
        ph_no_count = ph_no_count + 1;
        phone_number_cell = sheet.cell(row=count + ph_no_count, column=4)
        phone_number_cell.value = ph_nos

    blood_count = 0;
    for blood_grps in blood_grp:
        blood_count = blood_count + 1;
        blood_group_cell = sheet.cell(row=count + blood_count, column=5)
        blood_group_cell.value = blood_grps

    age_count = 0;
    for ages in age:
        age_count = age_count + 1;
        age_group_cell = sheet.cell(row=count + age_count, column=6)
        age_group_cell.value = ages

    location_count = 0;   
    for locations in location:
        location_count = location_count + 1;
        location_cell = sheet.cell(row=count + location_count, column=7)
        location_cell.value = locations
        

    file_open.close()


# excel
book = Workbook()
book.remove(book.active)
sheet = book.create_sheet()
# workbook = Workbook()
# sheet = workbook.create_sheet()
font = Font(bold=True)

# cells
student_name_cell = sheet.cell(row=1, column=1)
student_name_cell.value = "STUDENT NAME"
student_name_cell.font = font

class_cell = sheet.cell(row=1, column=2)
class_cell.value = "CLASS & SECTION"
class_cell.font = font

roll_number_cell = sheet.cell(row=1, column=3)
roll_number_cell.value = "ROLL NUMBER"
roll_number_cell.font = font

phone_number_cell = sheet.cell(row=1, column=4)
phone_number_cell.value = "PHONE NUMBER"
phone_number_cell.font = font

blood_group_cell = sheet.cell(row=1, column=5)
blood_group_cell.value = "BLOOD GROUP"
blood_group_cell.font = font

age_group_cell = sheet.cell(row=1, column=6)
age_group_cell.value = "AGE"
age_group_cell.font = font

location_cell = sheet.cell(row=1, column=7)
location_cell.value = "LOCATION"
location_cell.font = font


list_of_files = [os.path.basename(x) for x in
                 glob.glob('C:/Users/vsriniv3/Desktop/projects/python/krishna/krishna-master/*.txt')]
blood_group = input("Enter the Blood Group : ").upper()
location_input = input("Enter the location : ").lower()
for file in list_of_files:
    collect(sheet, file, blood_group, location_input)

# to save excel
book.save("results.xlsx")

print("\nPlease refer the Excel Sheet")
